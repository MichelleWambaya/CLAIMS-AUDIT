"""
Server-side Excel export (§7): multi-sheet workbook — summary + full
detail with flag reasons, duplicate group ID, days-from-first-visit, and
similarity score. Generated server-side (not in the browser) so it isn't
limited by browser memory on large result sets.

Summary sheet uses live formulas (COUNTIF/SUMIF) referencing the detail
sheet, not pre-computed Python totals, so the workbook stays correct if a
row is edited after export. Must be recalculated with LibreOffice after
writing — openpyxl never caches formula results (see xlsx skill notes).
"""
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .brand import COLOR_ORANGE, COLOR_BLACK, COLOR_WHITE

HEADER_FONT = Font(name="Arial", bold=True, color=COLOR_WHITE, size=11)
HEADER_FILL = PatternFill(start_color=COLOR_BLACK, end_color=COLOR_BLACK, fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
ACCENT_FONT = Font(name="Arial", bold=True, color=COLOR_ORANGE, size=12)

DETAIL_HEADERS = [
    "Row ID", "Flag Type", "Member ID", "Category", "Product / Diagnosis",
    "Amount", "Flag Reason", "Duplicate Group ID", "Days From First Visit",
    "Similarity Score", "Review Status",
]


def _style_header_row(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, ncols: int, min_width=12, max_width=40):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def build_workbook(
    session_name: str,
    flags: List[Dict[str, Any]],
    output_path: str,
):
    """
    flags: list of dicts with keys matching DETAIL_HEADERS (snake_case),
    e.g. {'row_id':.., 'flag_type':.., 'member_id':.., 'category':..,
    'product_or_diagnosis':.., 'amount':.., 'flag_reason':..,
    'duplicate_group_id':.., 'days_from_first_visit':..,
    'similarity_score':.., 'review_status':..}
    """
    wb = Workbook()

    detail_ws = wb.active
    detail_ws.title = "Flag Detail"
    for c, header in enumerate(DETAIL_HEADERS, start=1):
        detail_ws.cell(row=1, column=c, value=header)
    _style_header_row(detail_ws, 1, len(DETAIL_HEADERS))

    field_order = [
        "row_id", "flag_type", "member_id", "category", "product_or_diagnosis",
        "amount", "flag_reason", "duplicate_group_id", "days_from_first_visit",
        "similarity_score", "review_status",
    ]
    for r, flag in enumerate(flags, start=2):
        for c, field in enumerate(field_order, start=1):
            cell = detail_ws.cell(row=r, column=c, value=flag.get(field))
            cell.font = BODY_FONT
        if r % 2 == 0:
            for c in range(1, len(DETAIL_HEADERS) + 1):
                detail_ws.cell(row=r, column=c).fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )

    detail_ws.freeze_panes = "A2"
    _autosize(detail_ws, len(DETAIL_HEADERS))
    last_row = len(flags) + 1

    # --- Summary sheet, formula-driven against the detail sheet ---
    summary_ws = wb.create_sheet("Summary", 0)  # index 0 -> first tab
    summary_ws["A1"] = f"Claims Forensic Audit — {session_name}"
    summary_ws["A1"].font = ACCENT_FONT
    summary_ws.merge_cells("A1:C1")

    summary_ws["A3"] = "Flag Type"
    summary_ws["B3"] = "Count"
    _style_header_row(summary_ws, 3, 2)

    flag_types = [
        "item_duplicate", "claim_duplicate", "non_payable",
        "pricing_anomaly", "invalid_member_policy", "diagnosis_gap",
    ]
    for i, ft in enumerate(flag_types, start=4):
        summary_ws.cell(row=i, column=1, value=ft).font = BODY_FONT
        summary_ws.cell(row=i, column=2,
                         value=f'=COUNTIF(\'Flag Detail\'!B2:B{last_row},A{i})').font = BODY_FONT

    total_row = 4 + len(flag_types)
    summary_ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True)
    summary_ws.cell(row=total_row, column=2,
                     value=f"=SUM(B4:B{total_row - 1})").font = Font(name="Arial", bold=True)

    summary_ws.cell(row=total_row + 2, column=1, value="Confirmed").font = BODY_FONT
    summary_ws.cell(row=total_row + 2, column=2,
                     value=f'=COUNTIF(\'Flag Detail\'!K2:K{last_row},"confirmed")').font = BODY_FONT
    summary_ws.cell(row=total_row + 3, column=1, value="False Positive").font = BODY_FONT
    summary_ws.cell(row=total_row + 3, column=2,
                     value=f'=COUNTIF(\'Flag Detail\'!K2:K{last_row},"false_positive")').font = BODY_FONT
    summary_ws.cell(row=total_row + 4, column=1, value="Needs Follow-up").font = BODY_FONT
    summary_ws.cell(row=total_row + 4, column=2,
                     value=f'=COUNTIF(\'Flag Detail\'!K2:K{last_row},"needs_follow_up")').font = BODY_FONT

    _autosize(summary_ws, 3)

    wb.save(output_path)
    return output_path
