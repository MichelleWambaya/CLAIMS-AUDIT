"""
Streaming, row-by-row parsing for CSV and Excel so memory use doesn't scale
with file size. Per §4:
  - CSV: read incrementally, never `.read()` the whole file first.
  - Excel: openpyxl's `read_only=True` iterator mode streams rows without
    materializing the whole worksheet — this is the load-bearing choice,
    a plain `pandas.read_excel()` would defeat the point.
  - Sheet selection is explicit, never "just take the first sheet."

Each parser is a generator of dict rows keyed by the file's RAW headers.
Column mapping (rules/column_mapping.py) happens one layer up, after
per-file schema validation.
"""
import csv
import io
from typing import Iterator, Dict, Any, List, BinaryIO

import openpyxl


class SchemaIssue:
    def __init__(self, kind: str, detail: str):
        self.kind = kind        # 'missing_column' | 'unreadable_sheet' | 'wrong_header_row' | 'empty_file'
        self.detail = detail

    def to_dict(self):
        return {"kind": self.kind, "detail": self.detail}


def stream_csv_rows(byte_chunks: Iterator[bytes], encoding: str = "utf-8") -> Iterator[Dict[str, Any]]:
    """
    Wraps an iterator of raw byte chunks (e.g. from GraphClient.stream_download)
    as a line-based text stream, and yields one dict per row without ever
    holding the full file as a string.
    """
    def _line_iterator():
        buffer = b""
        for chunk in byte_chunks:
            buffer += chunk
            while b"\n" in buffer:
                line, buffer = buffer.split(b"\n", 1)
                yield line.decode(encoding, errors="replace")
        if buffer:
            yield buffer.decode(encoding, errors="replace")

    reader = csv.DictReader(_line_iterator())
    for row in reader:
        yield row


def list_workbook_sheets(file_obj: BinaryIO) -> List[str]:
    """Cheap sheet-name listing without loading cell data, for the
    per-workbook sheet-selection UI (§4: 'allow selecting which sheet')."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def stream_excel_rows(
    file_obj: BinaryIO,
    sheet_name: str,
    header_row_index: int = 0,
) -> "tuple[Iterator[Dict[str, Any]], List[SchemaIssue]]":
    """
    Streams rows from one sheet of a workbook using openpyxl's read-only
    iterator mode (does NOT load the whole sheet into memory).

    header_row_index: 0-based index of the header row, in case a cover
    sheet/banner pushes real headers down a row or two — configurable per
    file rather than assumed to be row 0.
    """
    issues: List[SchemaIssue] = []
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        issues.append(SchemaIssue("unreadable_sheet", f"Sheet '{sheet_name}' not found; available: {wb.sheetnames}"))
        wb.close()
        return iter([]), issues

    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)

    # Skip down to the configured header row.
    headers = None
    for i, row in enumerate(row_iter):
        if i == header_row_index:
            headers = [str(c).strip() if c is not None else "" for c in row]
            break

    if headers is None or all(h == "" for h in headers):
        issues.append(SchemaIssue("wrong_header_row", f"No usable header row found at index {header_row_index}"))
        wb.close()
        return iter([]), issues

    def _rows():
        for row in row_iter:
            if row is None or all(c is None for c in row):
                continue  # skip fully blank rows
            yield {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        wb.close()

    return _rows(), issues


def validate_required_columns(
    mapped_canonical_fields_present: set,
    required_fields: List[str],
) -> List[SchemaIssue]:
    missing = [f for f in required_fields if f not in mapped_canonical_fields_present]
    if missing:
        return [SchemaIssue("missing_column", f"Missing required field(s): {', '.join(missing)}")]
    return []
